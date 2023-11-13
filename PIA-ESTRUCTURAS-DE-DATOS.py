import sqlite3
from sqlite3 import Error
import sys
import datetime
import re
import csv
import openpyxl
try:
    with sqlite3.connect("Taller_Mecanico.db") as conn:
        mi_cursor = conn.cursor()

    
        mi_cursor.execute('''CREATE TABLE IF NOT EXISTS clientes (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            nombre TEXT NOT NULL,
                            rfc TEXT NOT NULL,
                            correo TEXT NOT NULL,
                            suspendido BOOLEAN DEFAULT 1
                        )''')

        mi_cursor.execute('''CREATE TABLE IF NOT EXISTS notas (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            cliente_id INTEGER,
                            folio INTEGER NOT NULL,
                            fecha timestamp NOT NULL,
                            monto_total REAL,
                            cancelada BOOLEAN DEFAULT 0,
                            FOREIGN KEY (cliente_id) REFERENCES clientes(id)
                        )''')

        mi_cursor.execute('''CREATE TABLE IF NOT EXISTS servicios (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            nombre TEXT NOT NULL,
                            costo REAL NOT NULL,
                            suspendido BOOLEAN DEFAULT 1
                            
                        )''')

        mi_cursor.execute('''CREATE TABLE IF NOT EXISTS detalles (
                            nota_id INTEGER,
                            servicio_id INTEGER,
                            PRIMARY KEY (nota_id, servicio_id),
                            FOREIGN KEY (nota_id) REFERENCES notas(id),
                            FOREIGN KEY (servicio_id) REFERENCES servicios(id)
                        )''')
except Error as e:
    print(e)
except Exception:
    print(f"Se produjo un problema:{sys.int_info()[0]}")
finally:
        
        conn.commit()

def validar_fecha(fecha):
    try:
        fecha_obj = datetime.datetime.strptime(fecha, "%d/%m/%Y")
        if fecha_obj <= datetime.datetime.now():
            return fecha_obj
        else:
            print("La fecha no puede ser posterior a la fecha actual.")
            return None
    except ValueError:
        print("Formato de fecha incorrecto. Debe ser mm/dd/aaaa.")
        return None
    


def agregar_cliente(nombre, rfc, correo):
    if not nombre or nombre.isspace():
        print("El nombre no puede estar vacío.")
        return

    if not re.match(r'^[A-ZÑ&]{3,4}\d{2}(0[1-9]|1[0-2])(0[1-9]|1\d|2\d|3[01])((H|M|h|m)(A|B|C|D|E|F|G|H|I|J|K|L|M|N|O|P|Q|R|S|T|U|V|W|X|Y|Z|NE|ne)[A-Z\d]{2}|[A-Z\d]{3})$', rfc):
        print("El RFC no es válido.")
        return

    if not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', correo):
        print("El correo electrónico no es válido.")
        return

    mi_cursor.execute("INSERT INTO clientes (nombre, rfc, correo) VALUES (?, ?, ?)",
                   (nombre, rfc, correo))

    conn.commit()
    


def agregar_servicio(nombre, costo):
    if not nombre or nombre.isspace():
        print("El nombre no puede estar vacío.")
        return

    if costo <= 0:
        print("El costo debe ser superior a 0.00.")
        return

    mi_cursor.execute("INSERT INTO servicios (nombre, costo) VALUES (?, ?)",
                   (nombre, costo))
    

    conn.commit()

def agregar_nota(cliente_id, servicios_ids):
    mi_cursor.execute("SELECT id FROM clientes WHERE id=? and suspendido = 1", (cliente_id,))
    cliente = mi_cursor.fetchone()

    if cliente is None:
        print("El cliente no está registrado.")
        return

    mi_cursor.execute("SELECT MAX(folio) FROM notas")
    max_folio = mi_cursor.fetchone()[0]
    folio = max_folio + 1 if max_folio is not None else 1

    while True:
        fecha_ingresada = input("Ingrese la fecha (dd/mm/aaaa): ")
        fecha_valida = validar_fecha(fecha_ingresada)
        if fecha_valida:
            break

    mi_cursor.execute("INSERT INTO notas (folio, cliente_id, fecha) VALUES (?, ?, ?)",
                   (folio, cliente_id, fecha_valida.strftime("%d/%m/%Y")))
    nota_id = mi_cursor.lastrowid

    monto_total = 0
    for servicio_id in servicios_ids:
        mi_cursor.execute("SELECT costo FROM servicios WHERE id=?", (servicio_id,))
        costo_servicio = mi_cursor.fetchone()[0]
        monto_total += costo_servicio

        mi_cursor.execute("INSERT INTO detalles (nota_id, servicio_id) VALUES (?, ?)", (nota_id, servicio_id))

    mi_cursor.execute("UPDATE notas SET monto_total=? WHERE id=?", (monto_total, nota_id))

    conn.commit()
    
def cancelar_nota(folio):
    
    mi_cursor.execute("SELECT * FROM notas WHERE folio=?", (folio,))
    nota = mi_cursor.fetchone()

    if nota is None:
        print("El folio indicado no existe.")
        return

   
    if nota[5]:  
        print("La nota ya ha sido cancelada.")
        print("")
        return

    print(f"Folio: {nota[2]}, Fecha: {nota[3]}, Monto total: {nota[4]}")

    
    confirmacion = input("¿Estás seguro de que quieres cancelar esta nota? (s/n): ")

    if confirmacion.lower() == "s":
        
        mi_cursor.execute("UPDATE notas SET cancelada=1 WHERE folio=?", (folio,))
        conn.commit()
        print("La nota ha sido cancelada.")
    else:
        print("La nota no ha sido cancelada.")

def recuperar_nota(folio):
    mi_cursor.execute("SELECT * FROM notas WHERE folio=?", (folio,))
    nota = mi_cursor.fetchone()

    if nota is None:
        print("")
        print("El folio indicado no existe.")
        return

    if not nota[5]:  
        print("")
        print("La nota no ha sido cancelada.")
        return
    print(f"Folio: {nota[2]}, Fecha: {nota[3]}, Monto total: {nota[4]}")

    confirmacion = input("¿Estás seguro de que quieres recuperar esta nota? (s/n): ")

    if confirmacion.lower() == "s":
        mi_cursor.execute("UPDATE notas SET cancelada=0 WHERE folio=?", (folio,))
        conn.commit()
        print("La nota ha sido recuperada.")
    else:
        print("La nota no ha sido recuperada.")
        
def consulta_por_periodo():
    fecha_inicial_ingresada = input("Ingrese la fecha inicial (dd/mm/aaaa) o deje en blanco para asumir 01/01/2000: ")
    if fecha_inicial_ingresada:
        fecha_inicial = validar_fecha(fecha_inicial_ingresada)
        if fecha_inicial is None:
            return
    else:
        fecha_inicial = datetime.datetime(2000, 1, 1)

    fecha_final_ingresada = input("Ingrese la fecha final (dd/mm/aaaa) o deje en blanco para asumir la fecha actual: ")
    if fecha_final_ingresada:
        fecha_final = validar_fecha(fecha_final_ingresada)
        if fecha_final is None:
            return
    else:
        fecha_final = datetime.datetime.now()

    if fecha_final < fecha_inicial:
        print("La fecha final debe ser igual o posterior a la fecha inicial.")
        return


    mi_cursor.execute("""
        SELECT notas.folio, notas.fecha, clientes.nombre
        FROM notas
        JOIN clientes ON notas.cliente_id = clientes.id
        WHERE notas.fecha BETWEEN ? AND ?
        AND notas.cancelada = 0  -- Agregar esta condición para notas no canceladas
        GROUP BY notas.id
        """, (fecha_inicial.strftime("%d/%m/%Y"), fecha_final.strftime("%d/%m/%Y")))
    notas = mi_cursor.fetchall()

    if notas:
        print('')
        print("Notas emitidas en el período especificado:")
        print('')
        print("*" * 50)
        print("{:<10} {:<15} {:<15}".format('Folio', 'Fecha', 'Nombre'))
        for nota in notas:
            print("{:<10} {:<15} {:<15}".format(nota[0], nota[1], nota[2]))
        print("*" * 50)
        print("")
    else:
        print("No hay notas emitidas para el período especificado.")
        
def consulta_por_folio():
    mi_cursor.execute("""
        SELECT notas.folio, notas.fecha, clientes.nombre
        FROM notas
        JOIN clientes ON notas.cliente_id = clientes.id
        WHERE notas.cancelada = 0
        ORDER BY notas.folio
        """)
    notas = mi_cursor.fetchall()

    print("Notas existentes:")
    for nota in notas:
        print(f"Folio: {nota[0]}, Fecha: {nota[1]}, Nombre del cliente: {nota[2]}")
    while True:
        folio = input("Ingrese el folio de la nota a consultar: ")
        
        if folio.strip()=="":
            print("FOLIO INVALIDO O NO EXISTE")
            print("")
            continue

        mi_cursor.execute("""
            SELECT notas.folio, notas.fecha, clientes.nombre, clientes.rfc, clientes.correo, GROUP_CONCAT(servicios.nombre, ', '), notas.monto_total
            FROM notas
            JOIN clientes ON notas.cliente_id = clientes.id
            JOIN detalles ON notas.id = detalles.nota_id
            JOIN servicios ON detalles.servicio_id = servicios.id
            WHERE notas.folio = ? AND notas.cancelada = 0
            GROUP BY notas.id
            """, (folio,))
        nota = mi_cursor.fetchone()

        if nota:
            
            print("")
            print('*' * 50)
            print(f"Folio: {nota[0]}")
            print(f"Fecha: {nota[1]}")
            print(f"Nombre del cliente: {nota[2]}")
            print(f"RFC del cliente: {nota[3]}")
            print(f"Correo del cliente: {nota[4]}")
            print(f"Servicios: {nota[5]}")
            print(f"Monto total: {nota[6]}")
            print('*' * 50)
            print("")
            break
        else:
            print("El folio indicado no existe o corresponde a una nota cancelada.")
            
            
def servicios_mas_prestados():
    while True:
        try:
            cantidad = int(input("Ingrese la cantidad de servicios más prestados a identificar: "))
            if cantidad <= 0:
                print("La cantidad debe ser un número positivo. Inténtalo de nuevo.")
                continue
            break
        except ValueError:
            print("Entrada inválida. Por favor, ingrese un número entero.")
    
    while True:
        fecha_inicial_ingresada = input("Ingrese la fecha inicial (dd/mm/aaaa): ")
        fecha_inicial = validar_fecha(fecha_inicial_ingresada)
        if fecha_inicial is None:
            print("Fecha inicial inválida. Inténtalo de nuevo.")
        else:
            break

    while True:
        fecha_final_ingresada = input("Ingrese la fecha final (dd/mm/aaaa): ")
        fecha_final = validar_fecha(fecha_final_ingresada)
        if fecha_final is None:
            print("Fecha final inválida. Inténtalo de nuevo.")
        elif fecha_final < fecha_inicial:
            print("La fecha final debe ser igual o posterior a la fecha inicial. Inténtalo de nuevo.")
        else:
            break
   
    mi_cursor.execute("""
        SELECT servicios.nombre, COUNT(*) as conteo
        FROM detalles
        JOIN notas ON detalles.nota_id = notas.id
        JOIN servicios ON detalles.servicio_id = servicios.id
        WHERE notas.fecha BETWEEN ? AND ?
        GROUP BY servicios.id
        ORDER BY conteo DESC
        LIMIT ?
        """, (fecha_inicial.strftime("%d/%m/%Y"), fecha_final.strftime("%d/%m/%Y"), cantidad))
    servicios = mi_cursor.fetchall()

    
    print("Servicios más prestados en el período especificado: ")
    print("{:<15} {:<10}".format('Servicio','Conteo'))
    print("*" * 30)
    for servicio in servicios:
        print("{:<15} {:<10}".format(servicio[0], servicio[1]))


def clientes_con_mas_notas():
    while True:
        try:
            cantidad = int(input("Ingrese la cantidad de clientes con más notas a identificar: "))
            if cantidad <= 0:
                print("La cantidad debe ser un número positivo. Inténtalo de nuevo.")
                continue
            break
        except ValueError:
            print("Entrada inválida. Por favor, ingrese un número entero.")
    
    while True:
        fecha_inicial_ingresada = input("Ingrese la fecha inicial (dd/mm/aaaa): ")
        fecha_inicial = validar_fecha(fecha_inicial_ingresada)
        if fecha_inicial is None:
            print("Fecha inicial inválida. Inténtalo de nuevo.")
        else:
            break

    while True:
        fecha_final_ingresada = input("Ingrese la fecha final (dd/mm/aaaa): ")
        fecha_final = validar_fecha(fecha_final_ingresada)
        if fecha_final is None:
            print("Fecha final inválida. Inténtalo de nuevo.")
        elif fecha_final < fecha_inicial:
            print("La fecha final debe ser igual o posterior a la fecha inicial. Inténtalo de nuevo.")
        else:
            break


    mi_cursor.execute("""
        SELECT clientes.nombre, COUNT(*) as conteo
        FROM notas
        JOIN clientes ON notas.cliente_id = clientes.id
        WHERE notas.fecha BETWEEN ? AND ?
        GROUP BY clientes.id
        ORDER BY conteo DESC
        LIMIT ?
        """, (fecha_inicial.strftime("%d/%m/%Y"), fecha_final.strftime("%d/%m/%Y"), cantidad))
    clientes = mi_cursor.fetchall()

    print('')
    print("Clientes con más notas en el período especificado:")
    print('*' * 40)
    print("{:<15} {:<10}".format('Cliente','Conteo'))
    print("*" * 40)
    for cliente in clientes:
        print("{:<15} {:<10}".format(cliente[0], cliente[1]))
        print('*' * 40)

def promedio_montos_notas():
    while True:
        fecha_inicial_ingresada = input("Ingrese la fecha inicial (dd/mm/aaaa): ")
        fecha_inicial = validar_fecha(fecha_inicial_ingresada)
        if fecha_inicial is None:
            print("Fecha inicial inválida. Inténtalo de nuevo.")
        else:
            break

    while True:
        fecha_final_ingresada = input("Ingrese la fecha final (dd/mm/aaaa): ")
        fecha_final = validar_fecha(fecha_final_ingresada)
        if fecha_final is None:
            print("Fecha final inválida. Inténtalo de nuevo.")
        elif fecha_final < fecha_inicial:
            print("La fecha final debe ser igual o posterior a la fecha inicial. Inténtalo de nuevo.")
        else:
            break

  
    mi_cursor.execute("""
        SELECT monto_total
        FROM notas
        WHERE fecha BETWEEN ? AND ?
        """, (fecha_inicial.strftime("%d/%m/%Y"), fecha_final.strftime("%d/%m/%Y")))
    montos = mi_cursor.fetchall()

   
    if montos:
        promedio = sum(monto[0] for monto in montos) / len(montos)
        promedio = round(promedio, 2)  
        print('')
        print(f"El promedio de los montos de las notas en el período especificado es: ${promedio} pesos")
        print('')
    else:
        print("No hay notas en el período especificado.")
            
            

def inicio():
    while True:
        print("")
        print("Taller Mecánico\n")
        print("MENU PRINCIPAL")
        print("1. Nota")
        print("2. Cliente")
        print("3. Servicio")
        print("4. Estadisticas")
        print("5. Salir\n")
        
        opcion_menu = input("Elige una opción: ")
        print("")
        
        if opcion_menu.strip()=="":
            print("OPCION INVALIDA")
            print("")
            continue

        elif opcion_menu =="1":
            while True:
                print("")
                print("Menu Notas")
                print("1. Registrar una nota")
                print("2. Cancelar una nota")
                print("3. Recuperar nota")
                print("4. Consultas y Reportes")
                print("5. Volver al menu principal")
                print("")
                opcion_nota = input("Elige una opción: ")
                print("")
                
                if opcion_nota.strip()=="":
                    print("OPCION INVALIDA")
                    print("")
                    continue
                elif opcion_nota == "1":
                    
                    try:
                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                            mi_cursor = conn.cursor()
                            mi_cursor.execute("SELECT id, nombre FROM clientes where suspendido = 1")
                            
                            registro_clientes = mi_cursor.fetchall()
                            
                            print("Clientes Registrados")
                            print("")
                            print("Claves\tNombre")
                            print("*" * 30)
                            if registro_clientes:
                                for clave, nombre in registro_clientes:
                                    print(f"{clave:6}\t{nombre}")
                            print("")
                                    
                    except Error as e:
                        print(e)
                    else:
                        if registro_clientes:
                            try:
                                while True:
                                    cliente_id = input("Ingresa la clave del cliente donde quiere registrar la nota: ")
                                    print("")
                                    if cliente_id.strip()=="":
                                        print("")
                                        print("CLAVE INVALIDA\n")
                                        continue
                                        
                                        
                                    else:
                                        valores_cliente = {"clave": cliente_id}
                                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                                            mi_cursor = conn.cursor()
                                            mi_cursor.execute("SELECT id, nombre, rfc, correo FROM clientes WHERE id = :clave and suspendido = 1", valores_cliente)
                                            llave = False
                                            registro = mi_cursor.fetchall()
                                            if registro:
                                                llave = True
                                            if llave==True:
                                                try:
                                                    
                                                    with sqlite3.connect("Taller_Mecanico.db") as conn:
                                                        mi_cursor = conn.cursor()
                                                        mi_cursor.execute("SELECT id, nombre, costo FROM servicios WHERE suspendido = 1")
                                                        
                                                        registro_datos_servicios = mi_cursor.fetchall()
                                                        if registro_datos_servicios:
                                                            print("CLAVE/\tNOMBRE\t/COSTO")
                                                            print("*" *30)
                                                            for clave, nombre, costo in registro_datos_servicios:
                                                                print(f"{clave:^6}{nombre}\t{costo}")
                                                                print("")
                                            
                                                            servicios_ids = input("Ingresa las claves de los servicios (separadas por espacios): ")
                                                            
                                                            valores_nota = {"servicio": servicios_ids}
                                                            with sqlite3.connect("Taller_Mecanico.db") as conn:
                                                                mi_cursor = conn.cursor()
                                                                mi_cursor.execute("SELECT id, nombre, costo FROM servicios WHERE id = :servicio and suspendido = 1", valores_nota)
                                                                registro = mi_cursor.fetchall()
                                                                if registro:
                                                                    agregar_nota(cliente_id, servicios_ids)
                                                                    print("")
                                                                    print("NOTA REGISTRADA")
                                                                    print("")
                                                                    break
                                                                else:
                                                                    print(f"No se encontro ningun servicio asociado a la clave: {servicios_ids}")
                                                        else:
                                                            print("* Aun no hay servicios registrdos *")
                                                            break
                                                except Error as e:
                                                    print(e)
                                            else:
                                                print(f"No se encontro el cliente con la clave {cliente_id}")
                                                print("")
                            except Error as e:
                                print(e)
                            except Exception:
                                print(f"Se produjo el siguiente error: {sys.int_info()[0]}")
                        
                        else:
                            print("* No se econtraron clientes *")
                            
  
                            
                            
                elif opcion_nota == "2":
                    folio = input("Ingrese el folio de la nota a cancelar: \n")
                    valores = {"folio":folio}
                    try:
                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                            mi_cursor = conn.cursor()
                            mi_cursor.execute("SELECT * FROM notas WHERE folio = :folio", valores)
                            llave2 = False
                            registro = mi_cursor.fetchall()
                            if registro:
                                llave2 = True
                                
                            else:
                                print(f"No se encontro una nota en el sistema con el siguiente folio: {folio}")
                                print("")
                                
                            if llave2==True:
                                cancelar_nota(folio)
                                print("")
                            
                    
                    except Error as e:
                        print(e)
                    finally:
                        conn.commit()
                        
                elif opcion_nota == "3":
                    folio = input("Ingresa el folio de la nota a recuperar: ")
                    
                    recuperar_nota(folio)
                    print("")
                    
                elif opcion_nota == "4":
                    while True:
                        print("1. Consulta por período")
                        print("2. Consulta por folio")
                        print("3. Regresar al Menu anterior")
                        opcion_consultas = input("Elige una opción: ")
                        print("")
                        if opcion_consultas.strip()=="":
                            print("OPCION INVALIDA")
                            print("")
                            continue
                        
                        elif opcion_consultas == "1":
                            consulta_por_periodo()
                        elif opcion_consultas == "2":
                            consulta_por_folio()
                                
                        elif opcion_consultas == "3":
                            break
                      
                elif opcion_nota=="5":
                    break
                
                
                           
        elif opcion_menu=="2":
            while True:
                print("")
                print("1. Agregar Cliente")
                print("2. Suspender clinete")
                print("3. Recuperar Cliente")
                print("4. Consultas y Reportes")
                print("5. Volver al Menu")
                print("")
                
                opcion_clientes = input("Eliga una opcion:\n")
                print("")
                
                if opcion_clientes.strip()=="":
                    print("OPCION INVALIDA\n")
                    continue
                elif opcion_clientes=="1":
                    nombre = input("Ingresa el nombre del cliente: ")
                    rfc = input("Ingresa el RFC del cliente: ")
                    correo = input("Ingresa el correo electrónico del cliente: ")
                    agregar_cliente(nombre, rfc, correo)
                    print("")
                    print("* CLIENTE REGISTRADO *")
                    print("")
                    
                elif opcion_clientes == "2":
                    while True:
                        try:
                            with sqlite3.connect("Taller_Mecanico.db") as conn:
                                mi_cursor = conn.cursor()
                                mi_cursor.execute("SELECT id, nombre FROM CLIENTES WHERE suspendido = 1")
                                registro = mi_cursor.fetchall()
                                print("Claves/Nombre")
                                print("*" * 20)
                                if registro:
                                    for clave, nombre, in registro:
                                        print(f"{clave:^7}{nombre}")
                                        print("*" * 20)
                                    cliente_suspender = input("Eliga la clave del cliente a suspender o 0 para salir: ")
                                    if cliente_suspender.strip()=="" or cliente_suspender.isalpha():
                                        print("")
                                        print("* DATO INVALIDO *")
                                        print("")
                                        continue
                              
                                    print("")
                                    if cliente_suspender == "0":
                                        break
                                    
                                    valores = {"cliente_suspendido": cliente_suspender}
                                    
                                    with sqlite3.connect("Taller_Mecanico.db") as conn:
                                        mi_cursor = conn.cursor()
                                        mi_cursor.execute("SELECT id, nombre, rfc, correo FROM clientes WHERE id = :cliente_suspendido", valores)
                                        registro = mi_cursor.fetchall()
                                        print("")
                                        print("Datos del Cliente a suspender\n")
                                        print("CLAVE/NOMBRE/RFC/CORREO")
                                        print("*" * 60)
                                        for clave, nombre, rfc, correo in registro:
                                                print(f"{clave:^6}{nombre}\t{rfc}\t{correo}")
                                                print("*" * 60)
                                                print("")
                                        confirmacion_suspender = input("Desea confirmar esta accion (S) o (N) para regresar al menu anterior: ")
                                        if confirmacion_suspender.strip()=="" or confirmacion_suspender.isnumeric():
                                            print("")
                                            print("* DATO INVALIDO *")
                                            print("")
                                            continue
                                        if confirmacion_suspender.upper()=="S":
                                            with sqlite3.connect("Taller_Mecanico.db") as conn:
                                                mi_cursor = conn.cursor()
                                                mi_cursor.execute("UPDATE clientes set suspendido = 0 WHERE id = :cliente_suspendido", valores)
                                                print("")
                                                print("* CLIENTE SUSPENDIDO *")
                                                break
                                        
                                        elif confirmacion_suspender == "N":
                                            break
                                        
                                        else:
                                            print("OPCION INVALIDA\n")
                                            continue
                                        
                                        
                                        
                                                
                                        
                            
                                            
                        except Error as e:
                            print(e)
                        
                        except Exception:
                            print(f"Se produjo el siguiente error: {sys.int_info()[0]}")
                            
                elif opcion_clientes == "3":
                    
                    print("* CLIENTES SUSPENDIDOS *")
                    with sqlite3.connect("Taller_Mecanico.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT id, nombre FROM clientes WHERE suspendido = 0")
                        registro_clientes_suspendidos = mi_cursor.fetchall()
                        print("CLAVE/NOMBRE")
                        print("*" * 20)
                        if registro_clientes_suspendidos:
                            for clave, nombre, in registro_clientes_suspendidos:
                                        print(f"{clave:^7}{nombre}")
                                        print("*" * 20)
                            print("")
                            while True:
                                cliente_recuperar = input("Seleccione la clave del cliente a recuperar o 0 para regresar: ")
                                if cliente_recuperar.strip()=="" and cliente_recuperar.isalpha():
                                    print("")
                                    print("* DATO INVALIDO *")
                                    print("")
                                    continue
                                    
                                if cliente_recuperar == "0":
                                    break
                                
                                if cliente_recuperar:
                                    valores_cliente_recuperar = {"clave_recuperar": cliente_recuperar}
                                    print("Datos del cliente a recuperar\n")
                                    with sqlite3.connect("Taller_Mecanico.db") as conn:
                                        mi_cursor = conn.cursor()
                                        mi_cursor.execute("SELECT id, nombre, rfc, correo FROM clientes WHERE suspendido = 0 and id = :clave_recuperar", valores_cliente_recuperar)
                                        datos_cliente = mi_cursor.fetchall()
                                        print("CLAVE NOMBRE \tRFC\t\tCORREO")
                                        print("*" * 60)
                                        for clave, nombre, rfc, correo in datos_cliente:
                                            print(f"{clave:^7}{nombre}\t{rfc}\t{correo}")
                                            print("*" * 60)
                                            
                                        confirmacion_recuperar = input("Desea confirmar la recuperacion (S) o (N) para regresar al menu anterior: ")
                                        if confirmacion_recuperar.strip()=="" or confirmacion_recuperar.isnumeric():
                                            print("")
                                            print("* DATO INVALIDO *")
                                            print("")
                                            continue
                                        if confirmacion_recuperar.upper() == "S":
                                            with sqlite3.connect("Taller_Mecanico.db") as conn:
                                                mi_cursor = conn.cursor()
                                                mi_cursor.execute("UPDATE clientes set suspendido = 1 WHERE id = :clave_recuperar", valores_cliente_recuperar)
                                                print("")
                                                print("* CLIENTE RECUPERADO *")
                                                break
                                        else:
                                            if confirmacion_recuperar == "N":
                                                break
                                                
                    
                elif opcion_clientes=="4":
                    while True:
                        print("1. Listado de clientes registrados")
                        print("2. Busqueda por clave")
                        print("3. Busqueda por nombre")
                        print("4. Volver al menu de clientes")
                        print("")
                        
                        opcion_clientes_2 = input("Eliga una opcion:\n")
                        print("")
                        
                        if opcion_clientes_2.strip()=="":
                            print("CAMPO VACIO")
                            print("")
                            continue
                            
                        elif opcion_clientes_2=="1":
                            while True:
                                print("")
                                print("1. Ordenado Por Clave")
                                print("2. Ordenado Por Nombre")
                                print("3. Volver al menu anterior")
                                print("")
                                opcion_clientes_3 = input("Eliga una opcion:\n")
                                if opcion_clientes_3.strip()=="":
                                    print("CAMPO VACIO")
                                    print("")
                                    continue
                                elif opcion_clientes_3=="1":
                                    try:
                                        fecha_actual = datetime.date.today()
                                        fecha_reporte = (f"{fecha_actual}.csv")
                                        
                                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                                            mi_cursor = conn.cursor()
                                            mi_cursor.execute("select id, nombre, rfc, correo from clientes where suspendido = 1 order by id;")
                                            
                                            registro_ordenado_id = mi_cursor.fetchall()
                                            print("Clientes Ordenados por ID")
                                            print("")
                                            print("Claves\tNombre\tRFC\t\tCorreo")
                                            print("*" * 60)
                                            if registro_ordenado_id:
                                                for clave, nombre, rfc, correo in registro_ordenado_id:
                                                    print(f"{clave:^6}{nombre}\t{rfc}\t{correo}")
                                                    print("")
                                            
                
                                    except Error as e:
                                        print(e)
                                        
                                    except Exception:
                                        print(f"Se produjo el siguiente error: {sys.int_info()[0]}")
                                        
                                    else: 
                                        while True:
                                            opcion_decision = input("Desea exportar estos datos a CSV o Excel?(S/N)\n")
                                            if opcion_decision.strip()=="" or opcion_decision.isnumeric():
                                                print("")
                                                print("* RESPUESTA INVALIDA *")
                                                print("")
                                                continue
                                            else:
                                            
                                                if opcion_decision.upper()=="S":
                                                    print("1. CSV")
                                                    print("2. Excel")
                                                    print("")
                                                    opcion_exportar = input("Eliga una opcion:\n")
                                                    if opcion_exportar.strip()=="" or opcion_exportar.isalpha():
                                                        print("")
                                                        print("* OPCION INVALIDA *")
                                                        continue
                                                    else:
                                                        
                                                        print("")
                                                        if opcion_exportar=="1":
                                                            encabezados = ["CLAVE", "NOMBRE", "RFC", "CORREO"]
                                                            with open(fecha_reporte,"w", newline="") as reporte:
                                                                grabador = csv.writer(reporte)
                                                                grabador.writerow(encabezados)
                                                                for datos in registro:
                                                                    grabador.writerow(datos)
                                                                print("DATOS EXPORTADOS EXITOSAMENTE\n")
                                                                break
                                                                
                                                        elif opcion_exportar=="2":
                                                            fecha_actual = datetime.date.today()
                                                            fecha_reporte_excel = (f"{fecha_actual}.xlsx")
                                                            exportar_excel = registro_ordenado_id
                                                            libro = openpyxl.Workbook()
                                                            hoja = libro["Sheet"]
                                                            encabezado = ["CLAVE", "NOMBRE", "RFC", "CORREO"]
                                                            hoja.append(encabezado)
                                                            for cliente in exportar_excel:
                                                                hoja.append(cliente)
                                                            hoja.title = "ID_ordenado"
                                                            libro.save(fecha_reporte_excel)
                                                            print("DATOS EXPORTADOS EXITOSAMENTE ")
                                                            break
                                                elif opcion_decision.upper()=="N":
                                                    break
                                                            
                                                    
                                elif opcion_clientes_3=="2":
                                    try:
                                        fecha_actual = datetime.date.today()
                                        fecha_reporte = (f"{fecha_actual}.csv")
                                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                                            mi_cursor = conn.cursor()
                                            mi_cursor.execute("select id, nombre, rfc, correo from clientes WHERE suspendido = 1 order by nombre;")
                                            
                                            registro= mi_cursor.fetchall()
                                            print("Clientes Ordenados por ID")
                                            print("")
                                            print("Claves\tNombre\tRFC\t\tCorreo")
                                            print("*" * 60)
                                            if registro:
                                                for clave, nombre, rfc, correo in registro:
                                                    print(f"{clave:^6}{nombre}\t{rfc}\t{correo}")
                                                    print("")
                                    except Error as e:
                                        print(e)
                                    except Exception:
                                        print(f"Se produjo el siguiente error: {sys.int_info()[0]}")
                                        
                                    else:
                                        while True:
                                            opcion_decision = input("Desea exportar estos datos a CSV o Excel?(S/N)\n")
                                            print("")
                                            if opcion_decision.strip()=="" or opcion_decision.isnumeric():
                                                print("")
                                                print("* RESPUESTA INVALIDA *")
                                                print("")
                                                continue
                                            else:
                                                if opcion_decision.upper()=="S":
                                                    print("1. CSV")
                                                    print("2. Excel")
                                                    print("")
                                                    opcion_exportar = input("Eliga una opcion:\n")
                                                    if opcion_exportar.strip()=="" or opcion_exportar.isalpha()():
                                                        print("* OPCION INVALIDA *")
                                                        print("")
                                                        continue
                                                    else:
                                                            
                                                        if opcion_exportar=="1":
                                                            encabezados = ["CLAVE", "NOMBRE", "RFC", "CORREO"]
                                                            with open(fecha_reporte,"w", newline="") as reporte:
                                                                grabador = csv.writer(reporte)
                                                                grabador.writerow(encabezados)
                                                                for datos in registro:
                                                                    grabador.writerow(datos)
                                                                print("DATOS EXPORTADOS EXISTOSAMENTE\n")
                                                                break
                                                                
                                                        elif opcion_exportar=="2":
                                                            fecha_actual = datetime.date.today()
                                                            fecha_reporte_excel = (f"{fecha_actual}.xlsx")
                                                            exportar_excel = registro
                                                            libro = openpyxl.Workbook()
                                                            hoja = libro["Sheet"]
                                                            encabezado = ["CLAVE", "NOMBRE", "RFC", "CORREO"]
                                                            hoja.append(encabezado)
                                                            for cliente in exportar_excel:
                                                                hoja.append(cliente)
                                                            hoja.title = "NOMBRE_ordenado"
                                                            libro.save(fecha_reporte_excel)
                                                            print("DATOS EXPORTADOS EXITOSAMENTE ")
                                                            break
                                                elif opcion_decision.upper()=="N":
                                                    break
                                                
                                        
                                            
                                        
                                elif opcion_clientes_3=="3":
                                    break
                                
                                
                        elif opcion_clientes_2=="2":
                            while True:
                                clave_buscar=input("Ingrese la clave del cliente:\n")
                                if clave_buscar.strip()=="" or clave_buscar.isalpha():
                                    print("")
                                    print("* DATO INVALIDO *")
                                    print("")
                                    continue
                                else:
                                    valores = {"clave": clave_buscar}
                                    try:
                                        fecha_actual = datetime.date.today()
                                        fecha_reporte = (f"{fecha_actual}.csv")
                                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                                            mi_cursor = conn.cursor()
                                            mi_cursor.execute("SELECT id, nombre, rfc, correo FROM clientes WHERE id = :clave and suspendido = 1", valores)
                                            
                                            registro = mi_cursor.fetchall()
                                            
                                            print("")
                                            print("Claves\tNombre\tRFC\t\tCorreo")
                                            print("*" * 60)
                                            if registro:
                                                for clave, nombre, rfc, correo in registro:
                                                    print(f"{clave:^6}{nombre}\t{rfc}\t{correo}")
                                                    print("")
                                                
                                            else:
                                                print(f"No se encontro ningun cliente con la clave: {clave_buscar} o se encuentra suspendido\n")
                                                break
                                        
                                    except Error as e:
                                        print(e)
                                    except Exception:
                                        print(f"Se produjo el siguiente error: {sys.int_info()[0]}")
                                        print("")
                                        
                                    else:
                                        while True:
                                            opcion_decision = input("Desea exportar estos datos a CSV o Excel?(S/N)\n")
                                            if opcion_decision.strip()=="" or opcion_decision.isnumeric():
                                                print("")
                                                print("* RESPUESTA INVALIDA *")
                                                print("")
                                                continue
                                            else:
                                                
                                                if opcion_decision.upper()=="S":
                                                    print("1. CSV")
                                                    print("2. Excel")
                                                    print("")
                                                    opcion_exportar = input("Eliga una opcion:\n")
                                                    if opcion_exportar.strip()=="" or opcion_exportar.isalpha():
                                                        print("")
                                                        print("* OPCION INVALIDA *")
                                                        print("")
                                                        continue
                                                    if opcion_exportar=="1":
                                                        encabezados = ["CLAVE", "NOMBRE", "RFC", "CORREO"]
                                                        with open(fecha_reporte,"w", newline="") as reporte:
                                                            grabador = csv.writer(reporte)
                                                            grabador.writerow(encabezados)
                                                            for datos in registro:
                                                                grabador.writerow(datos)
                                                            print("DATOS EXPORTADOS EXISTOSAMENTE\n")
                                                            break
                                                                    
                                                    elif opcion_exportar=="2":
                                                        fecha_actual = datetime.date.today()
                                                        fecha_reporte_excel = (f"{fecha_actual}.xlsx")
                                                        exportar_excel = registro
                                                        libro = openpyxl.Workbook()
                                                        hoja = libro["Sheet"]
                                                        encabezado = ["CLAVE", "NOMBRE", "RFC", "CORREO"]
                                                        hoja.append(encabezado)
                                                        for cliente in exportar_excel:
                                                            hoja.append(cliente)
                                                        hoja.title = "ID_cliente"
                                                        libro.save(fecha_reporte_excel)
                                                        print("* DATOS EXPORTADOS EXITOSAMENTE *") 
                                                        print("")
                                                        break
                                                        
                                                        
                                                elif opcion_decision.upper()=="N":
                                                    break
                                            
                               
                            
                                
                        elif opcion_clientes_2=="3":
                            while True:
                                nombre_buscar=input("Ingrese el nombre del cliente:\n")
                                if nombre_buscar.strip()=="" or nombre_buscar.isnumeric():
                                    print("")
                                    print("* DATO INVALIDO *")
                                    print("")
                                    continue
                                else:
                                    
                                    valores = {"nombre": nombre_buscar}
                                    try:
                                        fecha_actual = datetime.date.today()
                                        fecha_reporte = (f"{fecha_actual}.csv")
                                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                                            mi_cursor = conn.cursor()
                                            mi_cursor.execute("SELECT id, nombre, rfc, correo FROM clientes WHERE nombre = :nombre and suspendido = 1", valores)
                                            
                                            registro = mi_cursor.fetchall()
                                            
                                            print("")
                                            print("Claves\tNombre\tRFC\t\tCorreo")
                                            print("*" * 60)
                                            if registro:
                                                for clave, nombre, rfc, correo in registro:
                                                    print(f"{clave:^6}{nombre}\t{rfc}\t{correo}")
                                                    print("")
                                            else:
                                                print(f"No se encontro ningun registro con el nombre: {nombre_buscar} o se encunetra suspendido\n")
                                                break
                                    except Error as e:
                                        print(e)
                                    except Exception:
                                        print(f"Se produjo el siguiente error: {sys.int_info()[0]}")
                                        print("")
                                    else:
                                        opcion_decision = input("Desea exportar estos datos a CSV o Excel?(S/N)\n")
                                        if opcion_decision.strip()=="" or opcion_decision.isnumeric():
                                            print("")
                                            print("* RESPUESTA INVALIDA *")
                                            print("")
                                            continue
                                        else:
                                            if opcion_decision.upper()=="S":
                                                print("1. CSV")
                                                print("2. Excel")
                                                print("")
                                                opcion_exportar = input("Eliga una opcion:\n")
                                                if opcion_exportar.strip()=="" or opcion_exportar.isalpha():
                                                    print("")
                                                    print("* OPCION INVALIDA *")
                                                    print("")
                                                    continue
                                                else:
                                                    if opcion_exportar=="1":
                                                        encabezados = ["CLAVE", "NOMBRE", "RFC", "CORREO"]
                                                        with open(fecha_reporte,"w", newline="") as reporte:
                                                            grabador = csv.writer(reporte)
                                                            grabador.writerow(encabezados)
                                                            for datos in registro:
                                                                grabador.writerow(datos)
                                                            print("DATOS EXPORTADOS EXISTOSAMENTE\n")
                                                            break
                                                                    
                                                    elif opcion_exportar=="2":
                                                        fecha_actual = datetime.date.today()
                                                        fecha_reporte_excel = (f"{fecha_actual}.xlsx")
                                                        exportar_excel = registro
                                                        libro = openpyxl.Workbook()
                                                        hoja = libro["Sheet"]
                                                        encabezado = ["CLAVE", "NOMBRE", "RFC", "CORREO"]
                                                        hoja.append(encabezado)
                                                        for cliente in exportar_excel:
                                                            hoja.append(cliente)
                                                        hoja.title = "NOMBRE_ordenado"
                                                        libro.save(fecha_reporte_excel)
                                                        print("* DATOS EXPORTADOS EXITOSAMENTE *")
                                                        print("")
                                                        break
                                            elif opcion_decision.upper()=="N":
                                                break
                                        
                        elif opcion_clientes_2=="4":
                            break
                    
                elif opcion_clientes=="5":
                    break
            
        elif opcion_menu=="3":
            while True:
                print("")
                print("1. Agregar Servicio")
                print("2. Suspender un Servicio")
                print("3. Recuperar un Servicio")
                print("4. Consultas y Reportes")
                print("5. Volver al Menu Principal")
                print("")
                opcion_servcios = input("Eliga una opcion:\n")
                print("")
                if opcion_servcios=="1":
                    nombre = input("Ingresa el nombre del servicio: ")
                    costo = float(input("Ingresa el costo del servicio: "))
                    agregar_servicio(nombre, costo)
                    print("")
                    print("* SERVICIO AGREGADO *")
                elif opcion_servcios == "2":
                    while True:
                        try:
                            with sqlite3.connect("Taller_Mecanico.db") as conn:
                                mi_cursor = conn.cursor()
                                mi_cursor.execute("SELECT id, nombre FROM servicios WHERE suspendido = 1")
                                registro = mi_cursor.fetchall()
                                print("Claves/Nombre")
                                print("*" * 20)
                                if registro:
                                    for clave, nombre, in registro:
                                        print(f"{clave:^7}{nombre}")
                                        print("*" * 20)
                                        
                                    servicio_suspender = input("Eliga la clave del cliente a suspender o 0 para salir: ")
                                    if servicio_suspender.strip()=="":
                                        print("")
                                        print("* DATO INVALIDO *")
                                        print("")
                                        continue
                                    
                                    if servicio_suspender == "0":
                                        break
                                    else:
                                        valores = {"servicio_suspendido": servicio_suspender}
                                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                                            mi_cursor = conn.cursor()
                                            mi_cursor.execute("SELECT id, nombre, costo FROM servicios WHERE id = :servicio_suspendido", valores)
                                            registro = mi_cursor.fetchall()
                                            print("")
                                            print("Datos del Servicio a suspender\n")
                                            print("CLAVE/NOMBRE/COSTO")
                                            print("*" * 50)
                                            for clave, nombre, costo in registro:
                                                    print(f"{clave:^6} {nombre} {costo}")
                                                    print("*" * 50)
                                                    print("")
                                            confirmacion_suspender = input("Desea confirmar esta accion (S) o (N) para regresar al menu anterior: ")
                                            if confirmacion_suspender.strip()=="" or confirmacion_suspender.isnumeric():
                                                print("")
                                                print("* RESPUESTA INVALIDA *")
                                                print("")
                                                continue
                                            else:
                                                if confirmacion_suspender.upper()=="S":
                                                    with sqlite3.connect("Taller_Mecanico.db") as conn:
                                                        mi_cursor = conn.cursor()
                                                        mi_cursor.execute("UPDATE servicios set suspendido = 0 WHERE id = :servicio_suspendido", valores)
                                                        print("")
                                                        print("* SERVICIO SUSPENDIDO *")
                                                        break
                                                
                                                elif confirmacion_suspender == "N":
                                                    break
                                                
                                                else:
                                                    print("OPCION INVALIDA\n")
                                                    continue
                                            
                        except Error as e:
                            print(e)
                        
                        except Exception:
                            print(f"Se produjo el siguiente error: {sys.int_info()[0]}")
                    
                elif opcion_servcios == "3":
                    while True:
                        print("* SERVICIOS SUSPENDIDOS *")
                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                            mi_cursor = conn.cursor()
                            mi_cursor.execute("SELECT id, nombre FROM servicios WHERE suspendido = 0")
                            registro_servicios_suspendidos = mi_cursor.fetchall()
                            print("CLAVE/NOMBRE")
                            print("*" * 20)
                            if registro_servicios_suspendidos:
                                for clave, nombre, in registro_servicios_suspendidos:
                                            print(f"{clave:^7}{nombre}")
                                            print("*" * 20)
                                            
                                print("")
                                servicio_recuperar = input("Seleccione la clave del servicio a recuperar o 0 para regresar: ")
                                if servicio_recuperar.strip()=="":
                                    print("")
                                    print("* DATO INVALIDO *")
                                    print("")
                                    continue
                                else:
                                    if servicio_recuperar == "0":
                                        break
                                    
                                    if servicio_recuperar:
                                        valores_servicio_recuperar = {"clave_recuperar": servicio_recuperar}
                                        print("Datos del servicio a recuperar\n")
                                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                                            mi_cursor = conn.cursor()
                                            mi_cursor.execute("SELECT id, nombre, costo FROM servicios WHERE suspendido = 0 and id = :clave_recuperar", valores_servicio_recuperar)
                                            datos_servicio = mi_cursor.fetchall()
                                            print("CLAVE NOMBRE COSTO")
                                            print("*" * 60)
                                            for clave, nombre, costo in datos_servicio:
                                                print(f"{clave:^7}{nombre} {costo}")
                                                print("*" * 60)
                                                
                                            confirmacion_recuperar = input("Desea confirmar la recuperacion (S) o (N) para regresar al menu anterior: ")
                                            if confirmacion_recuperar.strip()=="" or confirmacion_recuperar.isnumeric():
                                                print("")
                                                print("* RESPUESTA INVALIDA *")
                                                print("")
                                                continue
                                            else:
                                                if confirmacion_recuperar.upper() == "S":
                                                    with sqlite3.connect("Taller_Mecanico.db") as conn:
                                                        mi_cursor = conn.cursor()
                                                        mi_cursor.execute("UPDATE servicios set suspendido = 1 WHERE id = :clave_recuperar", valores_servicio_recuperar)
                                                        print("")
                                                        print("* SERVICIO RECUPERADO *")
                                                        break
                                                else:
                                                    if confirmacion_recuperar == "N":
                                                        break
                    
                elif opcion_servcios== "4":
                    while True:
                        print("")
                        print("1. Busqueda por clave de servicio")
                        print("2. Busqueda por nombre de servicio")
                        print("3. Listado de Servicios")
                        print("4. Volver al Menu de Servicios")
                        print("")
                        opcion_servcios_2 = input("Eliga una opcion:\n")
                        if opcion_servcios_2 == "1":
                            print("")
                            busqueda_clave_servicios = input("Ingrese la clave del Servicio: \n")
                            valores_servicios = {"clave": busqueda_clave_servicios}
                            try:
                                with sqlite3.connect("Taller_Mecanico.db") as conn:
                                    mi_cursor = conn.cursor()
                                    mi_cursor.execute("SELECT id, nombre, costo FROM servicios WHERE id = :clave and suspendido = 1", valores_servicios)
                                    registro_servicios = mi_cursor.fetchall()
                                    print("")
                                    
                                    if registro_servicios:
                                        print("CLAVE/SERVICIO/\tCOSTO")
                                        print("*" *30)
                                        for clave, nombre, costo in registro_servicios:
                                            print(f"{clave:^6}{nombre}\t{costo}")
                                            print("")
                                    else:
                                        print(f"No se encontrp ningun servicio con la clave: {busqueda_clave_servicios} o se encuentra suspendido")
                            except Error as e:
                                print(e)
                            except Exception:
                                print(f"Se produjo el siguiente errro: {sys.int_info()[0]}")
                                    
                        elif opcion_servcios_2 == "2":
                            nombre_servicio = input("Ingrese el nombre del servicio:\n")
                            valores_nombre = {"nombre": nombre_servicio}
                            try:
                                with sqlite3.connect("Taller_Mecanico.db") as conn:
                                    mi_cursor = conn.cursor()
                                    mi_cursor.execute("SELECT id, nombre, costo FROM servicios WHERE nombre = :nombre and suspendido = 1", valores_nombre)
                                    registro_nombre = mi_cursor.fetchall()
                                    
                                    if registro_nombre:
                                        print("CLAVE/NOMBRE/\tCOSTO")
                                        print("*" *30)
                                        for clave, nombre, costo in registro_nombre:
                                            print(f"{clave:^6}{nombre}\t{costo}")
                                            print("")
                                    else:
                                        print(f"No se encontro ningun servicio con el nombre: {nombre_buscar} o se encuentra suspendido")
                            except Error as e:
                                print(e)
                            except Exception:
                                print(f"Se produjo el siguiente error: {sys.int_info()[0]}")

                        elif opcion_servcios_2 == "3":
                            while True:
                                print("")
                                print("1. Ordenado Por Clave")
                                print("2. Ordenado Por Nombre")
                                print("3. Volver al Menu Principal")
                                print("")
                                opcion_servcios_3 = int(input("Eliga una opcion:\n"))
                                
                                if opcion_servcios_3 == 1:
                                    fecha_actual = datetime.date.today()
                                    fecha_reporte = (f"{fecha_actual}.csv")
                                    try:
                                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                                            mi_cursor = conn.cursor()
                                            mi_cursor.execute("SELECT id, nombre, costo FROM servicios WHERE suspendido = 1 Order by id;")
                                            registro_servicios_orden_clave = mi_cursor.fetchall()
                                            if registro_servicios_orden_clave:
                                                print("CLAVE/NOMBRE/\tCOSTO")
                                                print("*" *30)
                                                for clave, nombre, costo in registro_servicios_orden_clave:
                                                    print(f"{clave:^6}{nombre}\t{costo}")
                                                    print("")
                                            else:
                                                print("* Aun no hay servicos *")
                                    except Error as e:
                                        print(e)
                                    except Exception:
                                        print(f"Se produjo el siguiente error: {sys.int_info()[0]}")
                                        
                                    else:
                                        opcion_decision = input("Desea exportar estos datos a CSV o Excel?(S/N)\n")
                                        if opcion_decision.upper()=="S":
                                            print("1. CSV")
                                            print("2. Excel")
                                            print("")
                                            opcion_exportar = int(input("Eliga una opcion:\n"))
                                            print("")
                                            if opcion_exportar==1:
                                                encabezados = ["CLAVE", "NOMBRE", "COSTO", ]
                                                with open(fecha_reporte,"w", newline="") as reporte:
                                                    grabador = csv.writer(reporte)
                                                    grabador.writerow(encabezados)
                                                    for datos in registro_servicios_orden_clave:
                                                        grabador.writerow(datos)
                                                    print("DATOS EXPORTADOS EXISTOSAMENTE\n")
                                        
                                            elif opcion_exportar==2:
                                                fecha_actual = datetime.date.today()
                                                fecha_reporte_excel = (f"{fecha_actual}.xlsx")
                                                exportar_excel = registro_servicios_orden_clave
                                                libro = openpyxl.Workbook()
                                                hoja = libro["Sheet"]
                                                encabezado = ["CLAVE", "NOMBRE", "COSTO"]
                                                hoja.append(encabezado)
                                                for servicio in exportar_excel:
                                                    hoja.append(servicio)
                                                hoja.title = "ID_servicios"
                                                libro.save(fecha_reporte_excel)
                                                print("* DATOS EXPORTADOS EXITOSAMENTE *") 
                                                print("")
                                                break
                                            
                                        elif opcion_decision.upper() == "N":
                                            continue
                                            
                                        
                                elif opcion_servcios_3 == 2:
                                    fecha_actual = datetime.date.today()
                                    fecha_reporte = (f"{fecha_actual}.csv")
                                    try:
                                        with sqlite3.connect("Taller_Mecanico.db") as conn:
                                            mi_cursor = conn.cursor()
                                            mi_cursor.execute("SELECT id, nombre, costo FROM servicios WHERE suspendido = 1 Order by nombre;")
                                            registro_servicios_orden_nombre = mi_cursor.fetchall()
                                            if registro_servicios_orden_nombre:
                                                print("CLAVE/NOMBRE/\tCOSTO")
                                                print("*"*30)
                                                for clave, nombre, costo in registro_servicios_orden_nombre:
                                                    print(f"{clave:^6}{nombre}\t{costo}")
                                                    print("")
                                    except Error as e:
                                        print(e)
                                    except Exception:
                                        print(f"Se produjo el siguiente error: {sys.int_info()[0]}")
                                        
                                    else:
                                        opcion_decision = input("Desea exportar estos datos a CSV o Excel?(S/N)\n")
                                        if opcion_decision.upper()=="S":
                                            print("1. CSV")
                                            print("2. Excel")
                                            print("")
                                            opcion_exportar = int(input("Eliga una opcion:\n"))
                                            print("")
                                            if opcion_exportar==1:
                                                encabezados = ["CLAVE", "NOMBRE", "COSTO", ]
                                                with open(fecha_reporte,"w", newline="") as reporte:
                                                    grabador = csv.writer(reporte)
                                                    grabador.writerow(encabezados)
                                                    for datos in registro_servicios_orden_nombre:
                                                        grabador.writerow(datos)
                                                    print("DATOS EXPORTADOS EXISTOSAMENTE\n")
                                                    
                                            elif opcion_exportar==2:
                                                fecha_actual = datetime.date.today()
                                                fecha_reporte_excel = (f"{fecha_actual}.xlsx")
                                                exportar_excel = registro_servicios_orden_nombre
                                                libro = openpyxl.Workbook()
                                                hoja = libro["Sheet"]
                                                encabezado = ["CLAVE", "NOMBRE", "COSTO"]
                                                hoja.append(encabezado)
                                                for servicio in exportar_excel:
                                                    hoja.append(servicio)
                                                hoja.title = "NOMBRE_servicios"
                                                libro.save(fecha_reporte_excel)
                                                print("* DATOS EXPORTADOS EXITOSAMENTE *") 
                                                print("")
                                                break
                                        elif opcion_decision.upper()=="N":
                                            continue
                                elif opcion_servcios_3 == 3:
                                    break
                                    
                                    
                        elif opcion_servcios_2 == "4":
                            break
                        
                        else:
                            if opcion_clientes_2.strip()=="" or opcion_clientes_2.isalpha():
                                print("")
                                print("* DATO INVALIDO *")
                                print("")
                                continue
                    
                elif opcion_servcios == "5":
                    break
                           
        elif opcion_menu=="4":
            print("1. Servicios mas prestados")
            print("2. Clientes con mas notas")
            print("3. Promedio de lo montos de las notas\n")
            opcion_estadistica = input("Eliga una opcion: ")
            
            if opcion_estadistica.strip()=="":
                print("* CAMPO VACIO *\n")
            
            elif opcion_estadistica == "1":
                servicios_mas_prestados()
                
            elif opcion_estadistica == "2":
                clientes_con_mas_notas()
                
            elif opcion_estadistica == "3":
                promedio_montos_notas()
            else:
                print("OPCION INVALIDA")
                
        elif opcion_menu=="5":
            print("")
            print("SALIENDO DEL SISTEMA...")
            break
        
if __name__ == "__main__":
    inicio()